"""
utils/export.py
Xuất báo cáo ra Excel và PDF.
"""
from __future__ import annotations
import os
from datetime import datetime


def _get_downloads_path(title: str, ext: str) -> str:
    """Trả về đường dẫn file trong thư mục Downloads với timestamp."""
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(downloads, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(downloads, f"{title}_{timestamp}.{ext}")


def _find_unicode_font() -> str:
    """
    Tìm và đăng ký font TTF hỗ trợ Unicode/tiếng Việt.
    Trả về tên font đã register (hoặc 'Helvetica' nếu không tìm được).
    """
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    candidates = [
        # Windows
        ("C:/Windows/Fonts/arial.ttf",    "Arial"),
        ("C:/Windows/Fonts/times.ttf",    "TimesNew"),
        ("C:/Windows/Fonts/calibri.ttf",  "Calibri"),
        # macOS
        ("/Library/Fonts/Arial.ttf",                                        "Arial"),
        ("/System/Library/Fonts/Supplemental/Arial.ttf",                    "Arial"),
        # Linux
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",                 "DejaVu"),
        ("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf", "Liberation"),
    ]

    for path, name in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont(name, path))
                return name
            except Exception:
                continue

    # Không tìm được → fallback Helvetica (mất dấu nhưng không crash)
    return "Helvetica"


# ══════════════════════════════════════════════════════════════════════════════
# XUẤT EXCEL
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(data: list[dict], headers: list[str], keys: list[str],
                 title: str = "BaoCao") -> str:
    """
    Xuất danh sách dict ra file .xlsx, lưu vào ~/Downloads.
    Trả về đường dẫn file đã tạo.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("Cần cài openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # ── Tiêu đề lớn (merge toàn bộ hàng 1) ──
    last_col_letter = chr(64 + len(headers))
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = title
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A2A4A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── Header cột (hàng 2) ──
    thin   = Side(style="thin", color="D3CECC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(bold=True, color="1A2A4A")        # FIX: navy trên nền vàng
        cell.fill      = PatternFill("solid", fgColor="C9A84C") # gold
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    ws.row_dimensions[2].height = 22

    # ── Dữ liệu ──
    for row_idx, record in enumerate(data, 3):
        fill_color = "FFFFFF" if row_idx % 2 == 1 else "F0EDE6"
        for col_idx, key in enumerate(keys, 1):
            value          = record.get(key, "")
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = Font(size=10)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
        ws.row_dimensions[row_idx].height = 18

    # ── Tự động độ rộng cột ──
    # Dùng get_column_letter thay vì col[0].column_letter
    # để tránh lỗi MergedCell không có thuộc tính column_letter
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    for col_idx, col in enumerate(ws.columns, 1):
        # Bỏ qua MergedCell — chỉ lấy ô thường để tính độ rộng
        max_len = max(
            (len(str(c.value or "")) for c in col if not isinstance(c, MergedCell)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # ── Freeze dòng header ──
    ws.freeze_panes = "A3"

    path = _get_downloads_path(title, "xlsx")
    wb.save(path)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# XUẤT PDF
# ══════════════════════════════════════════════════════════════════════════════
def export_pdf(data: list[dict], headers: list[str], keys: list[str],
               title: str = "BaoCao") -> str:
    """
    Xuất danh sách dict ra file .pdf, lưu vào ~/Downloads.
    Trả về đường dẫn file đã tạo.
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise ImportError("Cần cài reportlab: pip install reportlab")

    # FIX: đăng ký font Unicode để tiếng Việt hiển thị đúng dấu
    font_name = _find_unicode_font()

    path      = _get_downloads_path(title, "pdf")
    page_size = landscape(A4) if len(headers) > 5 else A4

    doc = SimpleDocTemplate(
        path, pagesize=page_size,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm,
    )

    NAVY = colors.HexColor("#1A2A4A")
    GOLD = colors.HexColor("#C9A84C")
    ALT  = colors.HexColor("#F0EDE6")
    GRAY = colors.HexColor("#6B7896")

    styles = getSampleStyleSheet()
    story  = []

    # ── Tiêu đề ──
    story.append(Paragraph(title, ParagraphStyle(
        "VTitle", parent=styles["Normal"],
        fontName=font_name, fontSize=16,
        textColor=NAVY, alignment=TA_CENTER, spaceAfter=4,
    )))
    story.append(Paragraph(
        f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y  %H:%M')}  |  Tổng: {len(data)} bản ghi",
        ParagraphStyle("VSub", parent=styles["Normal"],
                       fontName=font_name, fontSize=9,
                       textColor=GRAY, alignment=TA_CENTER, spaceAfter=10),
    ))
    story.append(Spacer(1, 0.4*cm))

    # ── Bảng dữ liệu ──
    table_data = [headers]
    for record in data:
        row = []
        for k in keys:
            val = record.get(k, "")
            if val is None:
                val = ""
            elif isinstance(val, float) and val == int(val):
                val = f"{int(val):,}"
            elif isinstance(val, int):
                val = f"{val:,}"
            row.append(str(val))
        table_data.append(row)

    col_w = (page_size[0] - 3*cm) / len(headers)
    t = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
    t.setStyle(TableStyle([
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0),  NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0), (-1, 0),  font_name),   # FIX
        ("FONTSIZE",      (0, 0), (-1, 0),  10),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("TOPPADDING",    (0, 0), (-1, 0),  8),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  8),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, GOLD),   # gạch vàng dưới header

        # Dữ liệu
        ("FONTNAME",      (0, 1), (-1, -1), font_name),   # FIX
        ("FONTSIZE",      (0, 1), (-1, -1), 9),
        ("ALIGN",         (0, 1), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),

        # Màu xen kẽ
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, ALT]),

        # Viền
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D3CECC")),
        ("BOX",           (0, 0), (-1, -1), 1,   NAVY),
    ]))
    story.append(t)

    # ── Footer ──
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "— Hệ thống Quản Lý Khách Sạn —",
        ParagraphStyle("VFooter", parent=styles["Normal"],
                       fontName=font_name, fontSize=8,
                       textColor=GRAY, alignment=TA_CENTER),
    ))

    doc.build(story)
    return path